import io
from sqlalchemy.orm import Session
from app.models.user import User, UserRole
from app.services.user_service import create_user


def _load_wb(file_bytes: bytes):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(file_bytes))


def process_student_excel(file_bytes: bytes, db: Session) -> dict:
    """
    Expected columns:
    student_name | student_email | class | section | parent_name | parent_email | parent_phone
    """
    wb = _load_wb(file_bytes)
    ws = wb.active
    created = 0
    skipped = 0
    errors: list[str] = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        try:
            s_name, s_email, class_name, section, p_name, p_email, p_phone = (
                str(v).strip() if v is not None else "" for v in row[:7]
            )
            if not s_email:
                skipped += 1
                continue

            existing = db.query(User).filter(User.email == s_email).first()
            if existing:
                skipped += 1
                continue

            student, _ = create_user(
                db, s_name, s_email, UserRole.student,
                class_name=class_name, section=section
            )

            if p_email:
                parent = db.query(User).filter(User.email == p_email).first()
                if not parent:
                    parent, _ = create_user(
                        db, p_name, p_email, UserRole.parent, phone=p_phone
                    )
                student.parent_id = parent.id

            created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


def process_teacher_excel(file_bytes: bytes, db: Session) -> dict:
    """
    Expected columns: name | email | phone
    """
    wb = _load_wb(file_bytes)
    ws = wb.active
    created = 0
    skipped = 0
    errors: list[str] = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row in enumerate(rows, start=2):
        try:
            name, email, phone = (
                str(v).strip() if v is not None else "" for v in row[:3]
            )
            if not email:
                skipped += 1
                continue
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                skipped += 1
                continue
            create_user(db, name, email, UserRole.teacher, phone=phone)
            created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
