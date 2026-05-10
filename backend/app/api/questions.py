import io
from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks, UploadFile, File
from fastapi.responses import Response
from openpyxl import load_workbook, Workbook
from sqlalchemy.orm import Session

from app.core.deps import get_db, get_current_user, require_admin
from app.core.processing_state import question_generating
from app.models import Question, Chapter, Subject, Grade
from app.models.content import ChapterContent, ContentType
from app.models.question import DifficultyLevel
from app.schemas.question import QuestionCreate, QuestionOut
from app.services.question_service import generate_questions_for_chapter

router = APIRouter()

_CORRECT_OPTION_MAP = {
    'a': 'A', 'b': 'B', 'c': 'C', 'd': 'D',
    '1': 'A', '2': 'B', '3': 'C', '4': 'D',
    'option_a': 'A', 'option_b': 'B', 'option_c': 'C', 'option_d': 'D',
}


@router.get("/", response_model=list[QuestionOut], dependencies=[Depends(get_current_user)])
def list_questions(
    chapter_id: str | None = Query(None),
    subject_id: str | None = Query(None),
    grade_id: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = (
        db.query(
            Question,
            Chapter.title.label("chapter_title"),
            Subject.name.label("subject_name"),
            Grade.standard.label("grade_standard"),
        )
        .outerjoin(Chapter, Question.chapter_id == Chapter.id)
        .outerjoin(Subject, Chapter.subject_id == Subject.id)
        .outerjoin(Grade, Subject.grade_id == Grade.id)
    )

    if chapter_id:
        query = query.filter(Question.chapter_id == chapter_id)
    if subject_id:
        query = query.filter(Subject.id == subject_id)
    if grade_id:
        query = query.filter(Grade.id == grade_id)

    results = query.all()

    return [
        QuestionOut(
            id=row.Question.id,
            question_text=row.Question.question_text,
            option_a=row.Question.option_a,
            option_b=row.Question.option_b,
            option_c=row.Question.option_c,
            option_d=row.Question.option_d,
            correct_option=row.Question.correct_option,
            difficulty=row.Question.difficulty,
            chapter_id=row.Question.chapter_id,
            chapter_title=row.chapter_title,
            subject_name=row.subject_name,
            grade_standard=row.grade_standard,
            created_at=row.Question.created_at,
            is_ai_generated=row.Question.created_by is None,
        )
        for row in results
    ]


@router.post("/", response_model=QuestionOut)
def create_question(
    data: QuestionCreate,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    chapter = db.query(Chapter).filter(Chapter.id == data.chapter_id).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found")

    question = Question(
        question_text=data.question_text,
        option_a=data.option_a,
        option_b=data.option_b,
        option_c=data.option_c,
        option_d=data.option_d,
        correct_option=data.correct_option,
        difficulty=data.difficulty,
        chapter_id=data.chapter_id,
        created_by=current_user.id,
    )
    db.add(question)
    db.commit()
    db.refresh(question)

    subject = db.query(Subject).filter(Subject.id == chapter.subject_id).first()
    grade = db.query(Grade).filter(Grade.id == subject.grade_id).first() if subject else None

    return QuestionOut(
        id=question.id,
        question_text=question.question_text,
        option_a=question.option_a,
        option_b=question.option_b,
        option_c=question.option_c,
        option_d=question.option_d,
        correct_option=question.correct_option,
        difficulty=question.difficulty,
        chapter_id=question.chapter_id,
        chapter_title=chapter.title,
        subject_name=subject.name if subject else None,
        grade_standard=grade.standard if grade else None,
        created_at=question.created_at,
        is_ai_generated=False,
    )


@router.get("/template")
def download_template(_=Depends(require_admin)):
    """Return a pre-formatted .xlsx template for bulk question upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(["question_text", "option_a", "option_b", "option_c", "option_d", "correct_option", "difficulty"])
    ws.append(["What is 2 + 2?", "3", "4", "5", "6", "B", "easy"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=question_template.xlsx"},
    )


@router.post("/upload/{chapter_id}")
async def upload_questions_excel(
    chapter_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    """Bulk-import questions from an .xlsx file. Valid rows are imported; invalid rows are reported."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx files are accepted")

    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large (max 5 MB)")

    try:
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid or corrupted Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel file is empty")

    raw_headers = [
        str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]
    ]
    required_cols = ["question_text", "option_a", "option_b", "option_c", "option_d", "correct_option", "difficulty"]
    missing = [c for c in required_cols if c not in raw_headers]
    if missing:
        raise HTTPException(400, f"Missing required columns: {', '.join(missing)}")

    col_idx = {c: raw_headers.index(c) for c in required_cols}

    chapter = db.query(Chapter).filter(Chapter.id == chapter_id).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found")

    def cell(row: tuple, col: str) -> str:
        i = col_idx[col]
        val = row[i] if i < len(row) else None
        return str(val).strip() if val is not None else ""

    errors: list[dict] = []
    to_insert: list[Question] = []

    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue

        question_text = cell(row, "question_text")
        option_a = cell(row, "option_a")
        option_b = cell(row, "option_b")
        option_c = cell(row, "option_c")
        option_d = cell(row, "option_d")
        raw_correct = cell(row, "correct_option")
        raw_difficulty = cell(row, "difficulty")

        row_errors: list[str] = []

        if len(question_text) < 3:
            row_errors.append("question_text is empty or too short")
        for name, val in [("option_a", option_a), ("option_b", option_b), ("option_c", option_c), ("option_d", option_d)]:
            if not val:
                row_errors.append(f"{name} is empty")

        correct_option = _CORRECT_OPTION_MAP.get(raw_correct.lower())
        if not correct_option:
            upper = raw_correct.upper()
            correct_option = upper if upper in {"A", "B", "C", "D"} else None
        if not correct_option:
            row_errors.append(f"correct_option '{raw_correct}' must be A, B, C, or D")

        difficulty = raw_difficulty.lower()
        if difficulty not in {"easy", "medium", "hard"}:
            row_errors.append(f"difficulty '{raw_difficulty}' must be easy, medium, or hard")

        if row_errors:
            errors.append({"row": row_num, "reason": "; ".join(row_errors)})
            continue

        to_insert.append(Question(
            question_text=question_text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            correct_option=correct_option,
            difficulty=DifficultyLevel(difficulty),
            chapter_id=chapter_id,
            created_by=current_user.id,
        ))

    for q in to_insert:
        db.add(q)
    if to_insert:
        db.commit()

    return {"imported": len(to_insert), "skipped": len(errors), "errors": errors}


@router.post(
    "/generate/{chapter_id}",
    dependencies=[Depends(require_admin)],
    status_code=202,
)
async def regenerate_questions(
    chapter_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """Manually trigger question generation for a chapter (e.g. after a failed auto-gen)."""
    if chapter_id in question_generating:
        raise HTTPException(409, "Questions are already being generated for this chapter")

    chapter = db.query(Chapter).filter(Chapter.id == chapter_id).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found")

    subject = db.query(Subject).filter(Subject.id == chapter.subject_id).first()
    grade = db.query(Grade).filter(Grade.id == subject.grade_id).first() if subject else None

    simplified_sections = (
        db.query(ChapterContent)
        .filter(
            ChapterContent.chapter_id == chapter_id,
            ChapterContent.content_type == ContentType.simplified_text,
            ChapterContent.is_ai_generated == True,  # noqa: E712
        )
        .order_by(ChapterContent.order_index)
        .all()
    )
    combined_text = "\n\n".join(s.text_content for s in simplified_sections if s.text_content)
    if not combined_text:
        raise HTTPException(400, "No AI-generated content found — run AI Processor first")

    background_tasks.add_task(
        generate_questions_for_chapter,
        chapter_id=chapter_id,
        simplified_text=combined_text,
        subject_name=subject.name if subject else "",
        grade_standard=grade.standard if grade else 0,
    )

    return {"chapter_id": chapter_id, "status": "queued"}


@router.put("/{question_id}", response_model=QuestionOut, dependencies=[Depends(require_admin)])
def update_question(question_id: str, data: QuestionCreate, db: Session = Depends(get_db)):
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(404, "Question not found")

    chapter = db.query(Chapter).filter(Chapter.id == data.chapter_id).first()
    if not chapter:
        raise HTTPException(404, "Chapter not found")

    question.question_text = data.question_text
    question.option_a = data.option_a
    question.option_b = data.option_b
    question.option_c = data.option_c
    question.option_d = data.option_d
    question.correct_option = data.correct_option
    question.difficulty = data.difficulty
    question.chapter_id = data.chapter_id
    # created_by is intentionally not touched — preserves AI vs manual distinction

    db.commit()
    db.refresh(question)

    subject = db.query(Subject).filter(Subject.id == chapter.subject_id).first()
    grade = db.query(Grade).filter(Grade.id == subject.grade_id).first() if subject else None

    return QuestionOut(
        id=question.id,
        question_text=question.question_text,
        option_a=question.option_a,
        option_b=question.option_b,
        option_c=question.option_c,
        option_d=question.option_d,
        correct_option=question.correct_option,
        difficulty=question.difficulty,
        chapter_id=question.chapter_id,
        chapter_title=chapter.title,
        subject_name=subject.name if subject else None,
        grade_standard=grade.standard if grade else None,
        created_at=question.created_at,
        is_ai_generated=question.created_by is None,
    )


@router.delete("/{question_id}", dependencies=[Depends(require_admin)])
def delete_question(question_id: str, db: Session = Depends(get_db)):
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(404, "Question not found")

    db.delete(question)
    db.commit()
    return {"message": "Question deleted"}
